#!/usr/bin/env python3
"""
Small helper: read an Excel file and export parsed rows to parsed_courses.json
and a small courses_data.js that the static front-end can consume as window.PARSED_COURSES.

Usage:
  python export_excel_to_json.py              # reads 课表.xlsx in repo root
  python export_excel_to_json.py myfile.xlsx  # reads provided file

This script uses openpyxl. Install with:
  pip install -r requirements.txt
"""
import sys
import json
from pathlib import Path
from openpyxl import load_workbook


def first_non_empty_row(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return row
    return None


def sheet_rows_as_dicts(sheet):
    # Try to detect headers from the first non-empty row
    header_row = None
    header_idx = 1
    for i in range(1, 6):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row = [str(c).strip() if c is not None else "" for c in row]
            header_idx = i
            break

    # Fallback to generic column names if headers look like numbers or empty
    if header_row is None:
        header_row = []

    if not any(h for h in header_row):
        # create generic headers A,B,C...
        max_col = sheet.max_column
        header_row = [f"col_{i+1}" for i in range(max_col)]

    rows = []
    for row in sheet.iter_rows(min_row=header_idx+1, values_only=True):
        # stop when encountering many empty rows at end
        if all((c is None or str(c).strip() == "") for c in row):
            continue
        obj = {}
        for i, h in enumerate(header_row):
            key = h if h else f"col_{i+1}"
            val = row[i] if i < len(row) else None
            # normalize simple types to str when needed
            if isinstance(val, str):
                obj[key] = val.strip()
            else:
                obj[key] = val
        rows.append(obj)
    return header_row, rows


def export(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True)
    out = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        headers, rows = sheet_rows_as_dicts(sheet)
        out[sheetname] = {
            "headers": headers,
            "rows": rows,
        }

    # write JSON
    with open("parsed_courses.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # write a small JS file the frontend can load
    js_content = "window.PARSED_COURSES = " + json.dumps(out, ensure_ascii=False, indent=2) + ";\n"
    with open("courses_data.js", "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"Wrote parsed_courses.json and courses_data.js from {workbook_path}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("课表.xlsx")
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(2)
    export(path)
